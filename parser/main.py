from bs4 import BeautifulSoup
import openpyxl
import time

def main():
     print("Открываю файл messages.html")
     print("Если файл не найден, то консоль закроется...")
     print
     time.sleep(2)
     with open("messages.html", encoding="utf-8") as fp:
        soup = BeautifulSoup(fp, "html.parser")
        mydivs = soup.find_all("div", class_='text')
        print("#############################################")

        divs = []

        
        for div in mydivs:
            divs.append(div)

        count = 1
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for i in divs:
            text = str(i)
            start = text.find("Конкурент:")
            end = text.find("Кличка собаки:")
            name = ""

            if start != -1:
                name = text[start + 11:end - 5]
            else:
                start = text.find("Конкурсант:")
                end = text.find("Кличка собаки:")
                if start != -1:
                    name = text[start + 11:end - 5]
                else:
                    print('Не найдено в этом элементе')


            ###########################################
            text = str(i)
            start = text.find("Кличка собаки:")
            end = text.find("Номер телефона")
            nameDog = ""

            if start != -1:
                nameDog = text[start + 15:end - 5]
            else:
                print("Не найдено в этом элементе")

            #############################################
            text = str(i)
            start = text.find("Номер телефона:")
            end = text.find("<br/>Доп")
            phone = ""

            if start != -1:
                phone = text[start + 16:end]
                
                start = phone.find("tel:")
                end = phone.find('">')

                if start != -1:
                    phone = phone[start + 4:end]


                print(name + "-" + nameDog + "-" + phone)
            else:
                print("Не найдено в этом элементе")
         
            worksheet.cell(row=count, column=1, value=name)
            worksheet.cell(row=count, column=2, value=nameDog)
            worksheet.cell(row=count, column=3, value=phone)

            count += 1
        
        workbook.save('MegaSobaka.xlsx')
        print("#############################################")
        print
        print("Успешно! Закрываю файлы и завершаю работу...")
        time.sleep(5)
    

if __name__ == "__main__":
     main()
